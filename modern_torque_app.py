#!/usr/bin/env python3
import os
import re
import json
import threading
import tempfile
import requests  # new import for API calls
import pandas as pd
import serial.tools.list_ports
import openai
from openpyxl import load_workbook, Workbook  # For reading and generating the template

# New import for Excel to PDF conversion using win32com
try:
    import win32com.client
except ImportError:
    win32com = None

# NEW: Import printing support from PyQt6
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QGridLayout, QLabel,
    QComboBox, QPushButton, QHeaderView,
    QStatusBar, QTabWidget, QTableWidget, QTableWidgetItem, QDialog,
    QFormLayout, QLineEdit, QDialogButtonBox, QHBoxLayout,
    QStackedWidget, QDoubleSpinBox, QMessageBox, QFileDialog,
    QDateEdit, QToolButton, QMenu, QApplication, QCheckBox
)
from PyQt6.QtGui import QAction, QClipboard, QImage
from PyQt6.QtCore import QThread, pyqtSignal, Qt, QDate, QTimer

from db_handler_local import (
    init_db, insert_default_torque_table_data,
    get_torque_table, insert_raw_data, insert_summary,
    add_torque_entry, update_torque_entry, delete_torque_entry,
    get_app_setting, set_app_setting
)
from serial_reader import read_from_serial, find_fits_in_selected_row
from openai_handler import perform_extraction_from_image

def convert_excel_to_pdf(excel_path: str, pdf_path: str):
    """
    Convert an Excel file to PDF using the Excel COM interface (pywin32).
    Opens in read-only mode, disables alerts, and ensures the workbook is closed cleanly.
    """
    if win32com is None:
        raise ImportError(
            "win32com.client module is required for Excel to PDF conversion. "
            "Please install pywin32 and run on Windows."
        )

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(os.path.abspath(excel_path), ReadOnly=1)
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
    finally:
        wb.Close(SaveChanges=0)
        excel.Quit()
        del wb
        del excel

class SerialReaderWorker(QThread):
    reading_signal = pyqtSignal(float, list)

    def __init__(self, port, selected_row):
        super().__init__()
        self.port = port
        self.selected_row = selected_row
        self.stop_event = threading.Event()

    def run(self):
        BAUD_RATE = 9600

        def callback(target_torque):
            print(f"[DEBUG] Serial callback received torque: {target_torque}")
            if self.stop_event.is_set():
                return
            fits = find_fits_in_selected_row(target_torque, self.selected_row)
            if fits:
                print(f"[DEBUG] torque {target_torque} fits in ranges: {fits}")
            else:
                print(f"[DEBUG] torque {target_torque} did NOT fit any range")
            self.reading_signal.emit(target_torque, fits)

        try:
            print(f"[DEBUG] Starting serial read on {self.port} at {BAUD_RATE} baud...")
            read_from_serial(self.port, BAUD_RATE, callback, self.stop_event)
        except Exception as e:
            print("[DEBUG] Error in serial reading:", e)

    def stop(self):
        print("[DEBUG] stop_event set. Stopping serial reading.")
        self.stop_event.set()

def calc_applied_torques(max_torque: float) -> list[float]:
    """
    Given a maximum torque, calculates typical applied torques at ~92%, ~58%, and ~33%.
    Rounds each to the nearest 10.
    """
    factors = [0.916, 0.583, 0.333]
    results = []
    for f in factors:
        raw_val = max_torque * f
        rounded = round(raw_val / 10) * 10
        results.append(rounded)
    return results

def calc_allowance_range(applied_val: float) -> str:
    """
    Returns a min-max allowance range string with a 4-6% tolerance.
    """
    tolerance = 0.06 if applied_val < 10 else 0.04
    low = applied_val * (1 - tolerance)
    high = applied_val * (1 + tolerance)
    return f"{round(low,1)} - {round(high,1)}"

def generate_filename(template: str, variables: dict) -> str:
    """
    Replaces placeholders (e.g. {{CustomerCompany}}) with actual values.
    """
    filename = template
    for key, value in variables.items():
        placeholder = "{{" + key + "}}"
        filename = filename.replace(placeholder, str(value))
    return filename

class TorqueEntryDialog(QDialog):
    def __init__(self, parent=None, entry_data=None):
        super().__init__(parent)
        self.setWindowTitle("Torque Entry")
        self.setMinimumWidth(300)
        self.entry_data = entry_data or {}
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout(self)

        self.max_torque_edit = QLineEdit(str(self.entry_data.get("max_torque", "")))
        self.unit_edit = QLineEdit(self.entry_data.get("unit", ""))
        self.type_edit = QLineEdit(self.entry_data.get("type", ""))
        self.applied_torq_edit = QLineEdit(self.entry_data.get("applied_torq", ""))
        self.allowance1_edit = QLineEdit(self.entry_data.get("allowance1", ""))
        self.allowance2_edit = QLineEdit(self.entry_data.get("allowance2", ""))
        self.allowance3_edit = QLineEdit(self.entry_data.get("allowance3", ""))

        layout.addRow("Max Torque:", self.max_torque_edit)
        layout.addRow("Unit:", self.unit_edit)
        layout.addRow("Type:", self.type_edit)
        layout.addRow("Applied Torque (JSON):", self.applied_torq_edit)
        layout.addRow("Allowance 1:", self.allowance1_edit)
        layout.addRow("Allowance 2:", self.allowance2_edit)
        layout.addRow("Allowance 3:", self.allowance3_edit)

        self.max_torque_edit.textChanged.connect(self.auto_fill_applied_from_max)
        self.applied_torq_edit.textChanged.connect(self.auto_fill_allowances_from_applied)

        self.button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)
        self.setLayout(layout)

    def auto_fill_applied_from_max(self):
        txt = self.max_torque_edit.text().strip()
        if not txt:
            return
        match = re.search(r"[\d\.]+", txt)
        if match:
            try:
                max_torque = float(match.group())
            except ValueError:
                return
        else:
            return
        applied_list = calc_applied_torques(max_torque)
        self.applied_torq_edit.blockSignals(True)
        self.applied_torq_edit.setText(json.dumps(applied_list))
        self.applied_torq_edit.blockSignals(False)
        self.auto_fill_allowances_from_applied()

    def auto_fill_allowances_from_applied(self):
        txt = self.applied_torq_edit.text().strip()
        if not txt:
            return
        try:
            arr = json.loads(txt)
            if not isinstance(arr, list):
                return
        except (ValueError, json.JSONDecodeError):
            return
        for i in range(3):
            val = arr[i] if i < len(arr) else 0
            rng = calc_allowance_range(val)
            if i == 0:
                self.allowance1_edit.setText(rng)
            elif i == 1:
                self.allowance2_edit.setText(rng)
            else:
                self.allowance3_edit.setText(rng)

    def get_data(self):
        return {
            "max_torque": self.max_torque_edit.text().strip(),
            "unit": self.unit_edit.text().strip(),
            "type": self.type_edit.text().strip(),
            "applied_torq": self.applied_torq_edit.text().strip(),
            "allowance1": self.allowance1_edit.text().strip(),
            "allowance2": self.allowance2_edit.text().strip(),
            "allowance3": self.allowance3_edit.text().strip()
        }

class ModernTorqueApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Torque Testing Application")
        self.setGeometry(100, 100, 950, 650)
        # Variables to store last exported file paths for printing
        self.last_exported_summary_path = None
        self.last_exported_envelope_path = None

        self.results_by_range = {}
        self.customer_info = {}
        self.serial_worker = None
        self.selected_row = None

        # Load OpenAI settings
        self.openai_api_key = get_app_setting("openai_api_key")
        self.openai_model = get_app_setting("openai_model") or "gpt-4-turbo"
        try:
            self.openai_temperature = float(get_app_setting("openai_temperature") or 0.7)
        except ValueError:
            self.openai_temperature = 0.7
        try:
            self.openai_top_p = float(get_app_setting("openai_top_p") or 1.0)
        except ValueError:
            self.openai_top_p = 1.0
        try:
            self.openai_presence_penalty = float(get_app_setting("openai_presence_penalty") or 0.0)
        except ValueError:
            self.openai_presence_penalty = 0.0
        try:
            self.openai_frequency_penalty = float(get_app_setting("openai_frequency_penalty") or 0.0)
        except ValueError:
            self.openai_frequency_penalty = 0.0

        self.show_extracted_data = (get_app_setting("show_extracted_data") or "false").lower() == "true"

        self.setStyleSheet(self.load_stylesheet())
        self.init_ui()

    def load_stylesheet(self):
        return """
        QMainWindow {
            background-color: #FAFAFA;
            font-family: "Segoe UI", "Helvetica Neue", sans-serif;
        }
        QDialog {
            background-color: #FFFFFF;
            color: #333;
        }
        QDialog QLabel {
            color: #333;
            font-size: 14px;
        }
        QDialog QLineEdit, QDateEdit {
            background-color: #FFFFFF;
            color: #333;
            border: 1px solid #ccc;
            border-radius: 4px;
            padding: 4px;
        }
        QDialogButtonBox {
            background-color: transparent;
        }
        QTabWidget::pane {
            border: none;
            background: #FFFFFF;
        }
        QTabBar::tab {
            background: #E0E0E0;
            color: #555;
            padding: 10px 20px;
            margin: 3px;
            border-radius: 6px;
        }
        QTabBar::tab:selected {
            background: #3498db;
            color: #FFFFFF;
            font-weight: bold;
        }
        QPushButton {
            background-color: #3498db;
            border: none;
            color: #FFFFFF;
            padding: 10px 20px;
            font-size: 14px;
            border-radius: 6px;
        }
        QPushButton:disabled {
            background-color: #95a5a6;
        }
        QPushButton:hover:!pressed {
            background-color: #2980b9;
        }
        QComboBox, QLineEdit, QDateEdit {
            padding: 6px;
            font-size: 14px;
            border: 1px solid #ccc;
            border-radius: 4px;
            background-color: #FFFFFF;
            color: #333;
        }
        QComboBox QAbstractItemView {
            background-color: #F0F0F0;
            color: #333;
            selection-background-color: #3498db;
            selection-color: #FFFFFF;
        }
        QTableWidget {
            background-color: #FFFFFF;
            border: 1px solid #ccc;
            color: #333;
        }
        QHeaderView::section {
            background-color: #E0E0E0;
            color: #333;
            padding: 6px;
            border: 1px solid #ccc;
        }
        QLabel {
            font-size: 14px;
            color: #333;
        }
        QToolButton {
            background-color: #3498db;
            border: none;
            color: #FFFFFF;
            padding: 10px 20px;
            font-size: 14px;
            border-radius: 6px;
            min-width: 220px;
        }
        QMenu {
            background-color: #3498db;
            border: none;
            color: #FFFFFF;
            font-size: 14px;
            border-radius: 6px;
        }
        QMenu::item {
            padding: 8px 16px;
        }
        QMenu::item:selected {
            background-color: #2980b9;
            color: #FFFFFF;
        }
        QCheckBox {
            color: #333;
            font-size: 14px;
        }
        QCheckBox::indicator {
            width: 16px;
            height: 16px;
            border: 1px solid #ccc;
            background-color: #fff;
        }
        QCheckBox::indicator:checked {
            background-color: #3498db;
            image: none;
        }
        QCheckBox::indicator:unchecked {
            background-color: #fff;
            image: none;
        }
        """

    def init_ui(self):
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)

        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)

        self.init_testing_tab()
        self.init_settings_tab()
        self.init_report_templates_tab()
        self.load_torque_table_data()

    # ------------------------------ TESTING TAB ------------------------------
    def init_testing_tab(self):
        self.testing_tab = QWidget()
        main_layout = QVBoxLayout(self.testing_tab)

        # Top Info Grid
        info_grid = QGridLayout()
        row = 0

        info_grid.addWidget(QLabel("Manufacturer:"), row, 0)
        self.manufacturer_edit = QLineEdit()
        info_grid.addWidget(self.manufacturer_edit, row, 1)

        info_grid.addWidget(QLabel("Serial Number:"), row, 2)
        self.serial_number_edit = QLineEdit()
        info_grid.addWidget(self.serial_number_edit, row, 3)

        row += 1
        info_grid.addWidget(QLabel("Model:"), row, 0)
        self.model_edit = QLineEdit()
        info_grid.addWidget(self.model_edit, row, 1)

        info_grid.addWidget(QLabel("Calibration Date:"), row, 2)
        self.calibration_date_edit = QDateEdit()
        self.calibration_date_edit.setDate(QDate.currentDate())
        self.calibration_date_edit.setCalendarPopup(True)
        info_grid.addWidget(self.calibration_date_edit, row, 3)

        row += 1
        info_grid.addWidget(QLabel("Max Torque:"), row, 0)
        self.max_torque_combo = QComboBox()
        info_grid.addWidget(self.max_torque_combo, row, 1)
        self.max_torque_combo.currentIndexChanged.connect(self.on_max_torque_combo_changed)

        info_grid.addWidget(QLabel("Calibration Due:"), row, 2)
        self.calibration_due_edit = QDateEdit()
        self.calibration_due_edit.setDate(QDate.currentDate().addYears(1))
        self.calibration_due_edit.setCalendarPopup(True)
        info_grid.addWidget(self.calibration_due_edit, row, 3)

        row += 1
        info_grid.addWidget(QLabel("Unit #:"), row, 0)
        self.unit_number_edit = QLineEdit()
        info_grid.addWidget(self.unit_number_edit, row, 1)

        info_grid.addWidget(QLabel("Customer/Company:"), row, 2)
        self.customer_edit = QLineEdit()
        info_grid.addWidget(self.customer_edit, row, 3)

        row += 1
        info_grid.addWidget(QLabel("Phone Number:"), row, 0)
        self.phone_edit = QLineEdit()
        info_grid.addWidget(self.phone_edit, row, 1)

        info_grid.addWidget(QLabel("Address:"), row, 2)
        self.address_edit = QLineEdit()
        info_grid.addWidget(self.address_edit, row, 3)

        row += 1
        info_grid.addWidget(QLabel("Serial Port:"), row, 0)
        self.port_combo = QComboBox()
        self.port_combo.addItems(self.get_serial_ports())
        info_grid.addWidget(self.port_combo, row, 1)

        # Live Torque label
        self.live_torque_label = QLabel("Live Torque: --")
        self.live_torque_label.setStyleSheet("font-size: 48px; padding: 5px;")
        info_grid.addWidget(self.live_torque_label, row, 3)

        main_layout.addLayout(info_grid)

        # Test Results Table
        self.torque_table = QTableWidget()
        self.torque_table.setColumnCount(7)
        self.torque_table.setHorizontalHeaderLabels([
            "Applied Torque", "Min - Max Allowance",
            "Test 1", "Test 2", "Test 3", "Test 4", "Test 5"
        ])
        self.torque_table.setRowCount(3)
        self.torque_table.setEditTriggers(QTableWidget.EditTrigger.AllEditTriggers)
        self.torque_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        main_layout.addWidget(self.torque_table)

        self.load_max_torque_dropdown()

        # Buttons row
        btn_layout = QHBoxLayout()
        self.start_btn = QPushButton("Begin Test")
        self.start_btn.clicked.connect(self.start_test)
        btn_layout.addWidget(self.start_btn)

        # Modified End Test button (updates calibration date and wipes data)
        self.stop_btn = QPushButton("End Test and Wipe Data")
        self.stop_btn.clicked.connect(self.stop_test)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_btn)

        # Import Customer Info (unchanged)
        self.upload_info_btn = QToolButton()
        self.upload_info_btn.setText("Import Customer Info")
        self.upload_info_btn.setMinimumWidth(220)
        self.upload_info_btn.setPopupMode(QToolButton.ToolButtonPopupMode.MenuButtonPopup)
        info_menu = QMenu()
        action_upload_file = QAction("Upload image from computer", self)
        action_clipboard = QAction("Upload image/screenshot from clipboard", self)
        action_webcam = QAction("Take image from web camera", self)
        action_link = QAction("Import info from link", self)
        info_menu.addAction(action_upload_file)
        info_menu.addAction(action_clipboard)
        info_menu.addAction(action_webcam)
        info_menu.addAction(action_link)
        self.upload_info_btn.setMenu(info_menu)
        action_upload_file.triggered.connect(self.upload_customer_info_from_file)
        action_clipboard.triggered.connect(self.upload_customer_info_from_clipboard)
        action_webcam.triggered.connect(self.upload_customer_info_from_webcam)
        action_link.triggered.connect(self.upload_customer_info_from_link)
        btn_layout.addWidget(self.upload_info_btn)

        # New Export/Print drop-down
        self.export_print_btn = QToolButton()
        self.export_print_btn.setText("Export / Print")
        self.export_print_btn.setMinimumWidth(220)
        self.export_print_btn.setPopupMode(QToolButton.ToolButtonPopupMode.MenuButtonPopup)
        export_menu = QMenu()
        export_summary_action = QAction("Export Summary", self)
        export_summary_action.triggered.connect(self.export_summary)
        export_menu.addAction(export_summary_action)
        export_envelope_action = QAction("Export Envelope", self)
        export_envelope_action.triggered.connect(self.export_envelope)
        export_menu.addAction(export_envelope_action)
        print_summary_action = QAction("Print Summary", self)
        print_summary_action.triggered.connect(self.print_summary)
        export_menu.addAction(print_summary_action)
        print_envelope_action = QAction("Print Envelope", self)
        print_envelope_action.triggered.connect(self.print_envelope)
        export_menu.addAction(print_envelope_action)
        self.export_print_btn.setMenu(export_menu)
        btn_layout.addWidget(self.export_print_btn)
        main_layout.addLayout(btn_layout)

        # Extracted Data Section
        self.extracted_data_label = QLabel("Extracted Data:")
        main_layout.addWidget(self.extracted_data_label)
        self.extracted_data_table = QTableWidget()
        self.extracted_data_table.setColumnCount(2)
        self.extracted_data_table.setHorizontalHeaderLabels(["Field", "Value"])
        self.extracted_data_table.verticalHeader().setVisible(False)
        self.extracted_data_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        main_layout.addWidget(self.extracted_data_table)
        self.extracted_data_label.setVisible(self.show_extracted_data)
        self.extracted_data_table.setVisible(self.show_extracted_data)

        self.testing_tab.setLayout(main_layout)
        self.tab_widget.addTab(self.testing_tab, "Torque Testing")

    def get_serial_ports(self):
        ports = serial.tools.list_ports.comports()
        return [p.device for p in ports]

    def load_max_torque_dropdown(self):
        self.max_torque_combo.clear()
        table_data = get_torque_table()
        for row in table_data:
            txt = f"{row['max_torque']} {row['unit']} - {row['type']}"
            self.max_torque_combo.addItem(txt, userData=row)
        if table_data:
            self.max_torque_combo.setCurrentIndex(0)
            self.selected_row = table_data[0]
            self.display_pre_test_rows()

    def on_max_torque_combo_changed(self, index):
        row_data = self.max_torque_combo.itemData(index)
        if row_data:
            self.selected_row = row_data
            self.display_pre_test_rows()
        else:
            self.selected_row = None
            self.clear_torque_table()

    def display_pre_test_rows(self):
        self.clear_torque_table()
        if not self.selected_row:
            return
        try:
            applied_arr = json.loads(self.selected_row.get("applied_torq", "[]"))
        except json.JSONDecodeError:
            applied_arr = [0, 0, 0]
        for i in range(3):
            allowance_key = self.selected_row.get(f"allowance{i+1}", "")
            applied_val = applied_arr[i] if i < len(applied_arr) else 0
            self.torque_table.setItem(i, 0, QTableWidgetItem(str(applied_val)))
            self.torque_table.setItem(i, 1, QTableWidgetItem(allowance_key))
            for c in range(2, 7):
                self.torque_table.setItem(i, c, QTableWidgetItem(""))
        self.results_by_range = {}

    def clear_torque_table(self):
        for r in range(self.torque_table.rowCount()):
            for c in range(self.torque_table.columnCount()):
                self.torque_table.setItem(r, c, QTableWidgetItem(""))

    def start_test(self):
        if not self.selected_row:
            QMessageBox.warning(self, "Warning", "No torque entry selected.")
            return
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.statusBar.showMessage("Test in progress...")
        port = self.port_combo.currentText()
        if not port:
            QMessageBox.warning(self, "Warning", "No serial port selected.")
            return
        self.serial_worker = SerialReaderWorker(port, self.selected_row)
        self.serial_worker.reading_signal.connect(self.process_reading)
        self.serial_worker.start()

    def stop_test(self):
        if self.serial_worker:
            self.serial_worker.stop()
            self.serial_worker.wait(2000)
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.statusBar.showMessage("Test ended.")
        # Wipe test data and update calibration date
        self.results_by_range.clear()
        self.clear_torque_table()
        self.port_combo.clear()
        self.port_combo.addItems(self.get_serial_ports())
        self.calibration_date_edit.setDate(QDate.currentDate())
        QMessageBox.information(self, "Test Completed", "Test ended and data wiped.")
        print("[DEBUG] Test stopped. Results wiped.")

    def process_reading(self, target_torque, fits):
        self.live_torque_label.setText(f"Live Torque: {target_torque}")
        if fits:
            self.live_torque_label.setStyleSheet("background-color: green; color: white; font-size: 48px; padding: 5px;")
        else:
            self.live_torque_label.setStyleSheet("background-color: red; color: white; font-size: 48px; padding: 5px;")
        for fit in fits:
            allowance_key = fit.get('range_str', "")
            current_results = self.results_by_range.get(allowance_key, [])
            if len(current_results) < 5:
                insert_raw_data(
                    target_torque, self.selected_row["id"],
                    f"allowance{fit.get('allowance_index', '')}",
                    allowance_key
                )
                current_results.append(target_torque)
                self.results_by_range[allowance_key] = current_results
        self.update_summary_table()

    def update_summary_table(self):
        for row_idx in range(self.torque_table.rowCount()):
            allow_item = self.torque_table.item(row_idx, 1)
            if allow_item:
                allow_key = allow_item.text().strip()
                test_vals = self.results_by_range.get(allow_key, [])
                for col_idx in range(2, 7):
                    val_index = col_idx - 2
                    if val_index < len(test_vals):
                        self.torque_table.setItem(row_idx, col_idx, QTableWidgetItem(str(test_vals[val_index])))
                    else:
                        self.torque_table.setItem(row_idx, col_idx, QTableWidgetItem(""))

    # -------------------- CUSTOMER INFO IMPORTING --------------------
    def upload_customer_info_from_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Image", "",
            "Image Files (*.png *.jpg *.jpeg *.bmp);;All Files (*)"
        )
        if not file_path:
            return
        if not self.openai_api_key:
            QMessageBox.critical(self, "Error", "OpenAI API Key not set.")
            return
        extracted_data = self.extract_torque_data(file_path)
        if not extracted_data:
            QMessageBox.warning(self, "Extraction Failed", "No data extracted or an error occurred.")
            return
        self.update_extracted_data_table(extracted_data)

    def upload_customer_info_from_clipboard(self):
        clipboard = QApplication.clipboard()
        image = clipboard.image()
        if image.isNull():
            QMessageBox.warning(self, "Clipboard Empty", "No image found in clipboard.")
            return
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        temp_file.close()
        if not image.save(temp_file.name, "PNG"):
            QMessageBox.critical(self, "Error", "Failed to save clipboard image.")
            return
        if not self.openai_api_key:
            QMessageBox.critical(self, "Error", "OpenAI API Key not set.")
            return
        extracted_data = self.extract_torque_data(temp_file.name)
        if not extracted_data:
            QMessageBox.warning(self, "Extraction Failed", "No data extracted or an error occurred.")
            return
        self.update_extracted_data_table(extracted_data)

    def upload_customer_info_from_webcam(self):
        try:
            import cv2
        except ImportError:
            QMessageBox.critical(self, "Error", "OpenCV is not installed. Please install opencv-python.")
            return
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            QMessageBox.critical(self, "Error", "Could not open web camera.")
            return
        cv2.namedWindow("Webcam - Press Space to Capture", cv2.WINDOW_NORMAL)
        while True:
            ret, frame = cap.read()
            if not ret:
                QMessageBox.critical(self, "Error", "Failed to capture image from web camera.")
                cap.release()
                cv2.destroyAllWindows()
                return
            cv2.imshow("Webcam - Press Space to Capture", frame)
            key = cv2.waitKey(1) & 0xFF
            if key == 32:
                break
            elif key == 27:
                cap.release()
                cv2.destroyAllWindows()
                return
        cap.release()
        cv2.destroyAllWindows()
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        temp_file.close()
        cv2.imwrite(temp_file.name, frame)
        if not self.openai_api_key:
            QMessageBox.critical(self, "Error", "OpenAI API Key not set.")
            return
        extracted_data = self.extract_torque_data(temp_file.name)
        if not extracted_data:
            QMessageBox.warning(self, "Extraction Failed", "No data extracted or an error occurred.")
            return
        self.update_extracted_data_table(extracted_data)

    def upload_customer_info_from_link(self):
        clipboard = QApplication.clipboard()
        text = clipboard.text().strip()
        if not text:
            QMessageBox.warning(self, "Clipboard Empty", "No text found in clipboard.")
            return
        match = re.search(r'(\d+)(?!.*\d)', text)
        if not match:
            QMessageBox.warning(self, "Invalid Link", "No numeric ID found in the clipboard text.")
            return
        line_item_id = match.group(1)
        token = get_app_setting("laravel_api_token")
        if not token:
            QMessageBox.critical(self, "Error", "Laravel API token not set in settings.")
            return
        laravel_url = get_app_setting("laravel_app_url")
        if not laravel_url:
            QMessageBox.critical(self, "Error", "Laravel app URL not set in settings.")
            return
        line_item_response = self.get_line_item_from_api(line_item_id, token, laravel_url)
        if not line_item_response:
            return
        line_item_data = line_item_response.get("data")
        if not line_item_data:
            QMessageBox.warning(self, "API Error", "No 'data' key found in line item response.")
            return
        company_asset = line_item_data.get("company_asset", {})
        self.unit_number_edit.setText(company_asset.get("unit_number", ""))
        self.manufacturer_edit.setText(company_asset.get("make", ""))
        self.model_edit.setText(company_asset.get("model", ""))
        self.serial_number_edit.setText(company_asset.get("serial_number", ""))
        asset_info = company_asset.get("additional_info_fields", {})
        max_torque_str = str(asset_info.get("max_torque", "")).strip()
        torque_unit_str = str(asset_info.get("torque_unit", "")).strip()
        if max_torque_str:
            try:
                extracted_val = float(max_torque_str)
                self.auto_select_max_torque(extracted_val, torque_unit_str)
            except ValueError:
                pass
        company_id = company_asset.get("company_id")
        if company_id:
            company_response = self.get_company_info_from_api(company_id, token, laravel_url)
            if company_response:
                company_data = company_response.get("data", company_response)
                self.customer_edit.setText(company_data.get("name", ""))
                self.phone_edit.setText(company_data.get("phone", ""))
        QMessageBox.information(self, "Success", "Customer info imported from API.")

    def get_line_item_from_api(self, line_item_id, token, base_url):
        url = f"{base_url}/api/line-items/{line_item_id}"
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json"
        }
        try:
            response = requests.get(url, headers=headers)
        except Exception as e:
            QMessageBox.critical(self, "API Error", f"Error during line-item request: {e}")
            return None
        if response.status_code == 200:
            try:
                return response.json()
            except Exception as e:
                QMessageBox.critical(self, "API Error", f"Error parsing JSON response: {e}")
                return None
        else:
            QMessageBox.warning(self, "API Error", f"Line-item request failed: {response.status_code} {response.text}")
            return None

    def get_company_info_from_api(self, company_id, token, base_url):
        url = f"{base_url}/api/companies/{company_id}"
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json"
        }
        try:
            response = requests.get(url, headers=headers)
        except Exception as e:
            QMessageBox.critical(self, "API Error", f"Error during company request: {e}")
            return None
        if response.status_code == 200:
            try:
                return response.json()
            except Exception as e:
                QMessageBox.critical(self, "API Error", f"Error parsing JSON response for company: {e}")
                return None
        else:
            QMessageBox.warning(self, "API Error", f"Company request failed: {response.status_code} {response.text}")
            return None

    def extract_torque_data(self, image_path: str) -> dict:
        return perform_extraction_from_image(image_path, self.openai_api_key, self.openai_model)

    def update_extracted_data_table(self, data: dict):
        self.manufacturer_edit.setText(data.get("manufacturer", ""))
        self.model_edit.setText(data.get("model", ""))
        self.unit_number_edit.setText(data.get("unit", ""))
        self.serial_number_edit.setText(data.get("serial", ""))
        self.customer_edit.setText(data.get("customer", ""))
        self.phone_edit.setText(data.get("phone", ""))
        self.address_edit.setText(data.get("address", ""))
        max_torque_raw = data.get("max_torque", "")
        torque_unit_str = str(data.get("torque_unit", "")).strip()
        max_torque_str = str(max_torque_raw).strip()
        if max_torque_str:
            try:
                extracted_val = float(max_torque_str)
                self.auto_select_max_torque(extracted_val, torque_unit_str)
            except ValueError:
                pass
        fields = [
            ("Manufacturer", data.get("manufacturer", "")),
            ("Model", data.get("model", "")),
            ("Unit #", data.get("unit", "")),
            ("Serial Number", data.get("serial", "")),
            ("Customer/Company", data.get("customer", "")),
            ("Phone Number", data.get("phone", "")),
            ("Address", data.get("address", "")),
            ("Max Torque", max_torque_str),
            ("Torque Unit", torque_unit_str)
        ]
        self.extracted_data_table.setRowCount(len(fields))
        for i, (field, value) in enumerate(fields):
            self.extracted_data_table.setItem(i, 0, QTableWidgetItem(field))
            self.extracted_data_table.setItem(i, 1, QTableWidgetItem(value))

    def auto_select_max_torque(self, extracted_val: float, extracted_unit: str):
        ft_lb_synonyms_str = get_app_setting("synonyms_ft_lb") or "ft/lb,ft-lb,ft.lb,ft lb,ft/lbs,ft-lbs,ft.lbs,ft lbs"
        in_lb_synonyms_str = get_app_setting("synonyms_in_lb") or "in/lb,in-lb,in.lb,in lb,in/lbs,in-lbs,in.lbs,in lbs"
        nm_synonyms_str = get_app_setting("synonyms_nm") or "nm,n.m,n*m,nm.,n.m."
        FT_LB_SYNONYMS = set(s.strip() for s in ft_lb_synonyms_str.split(",") if s.strip())
        IN_LB_SYNONYMS = set(s.strip() for s in in_lb_synonyms_str.split(",") if s.strip())
        NM_SYNONYMS = set(s.strip() for s in nm_synonyms_str.split(",") if s.strip())
        def ftlb_to_nm(val):
            return val * 1.35582
        def inlb_to_nm(val):
            return val * 0.113
        extracted_unit_lower = extracted_unit.lower().strip()
        if extracted_unit_lower in FT_LB_SYNONYMS:
            extracted_val_nm = ftlb_to_nm(extracted_val)
        elif extracted_unit_lower in IN_LB_SYNONYMS:
            extracted_val_nm = inlb_to_nm(extracted_val)
        elif extracted_unit_lower in NM_SYNONYMS:
            extracted_val_nm = extracted_val
        else:
            extracted_val_nm = extracted_val
        table_data = get_torque_table()
        tolerance_base = max(extracted_val_nm * 0.10, 2.0)
        for i, row in enumerate(table_data):
            db_torque = row["max_torque"]
            db_unit_lower = row["unit"].lower().strip()
            if db_unit_lower in FT_LB_SYNONYMS:
                db_torque_nm = ftlb_to_nm(db_torque)
            elif db_unit_lower in IN_LB_SYNONYMS:
                db_torque_nm = inlb_to_nm(db_torque)
            elif db_unit_lower in NM_SYNONYMS:
                db_torque_nm = db_torque
            else:
                db_torque_nm = db_torque
            if abs(db_torque_nm - extracted_val_nm) <= tolerance_base:
                self.max_torque_combo.setCurrentIndex(i)
                self.selected_row = row
                self.display_pre_test_rows()
                return

    # ------------------------------ EXPORTING SUMMARY ------------------------------
    def export_summary(self):
        row_count = self.torque_table.rowCount()
        headers = ["Applied Torque", "Min - Max Allowance", "Test 1", "Test 2", "Test 3", "Test 4", "Test 5"]
        summary_data = []
        for r in range(row_count):
            row_dict = {}
            for c in range(self.torque_table.columnCount()):
                item = self.torque_table.item(r, c)
                row_dict[headers[c]] = item.text() if item else ""
            summary_data.append(row_dict)
        extra_info = {
            "Manufacturer": self.manufacturer_edit.text(),
            "Serial Number": self.serial_number_edit.text(),
            "Model": self.model_edit.text(),
            "Calibration Date": self.calibration_date_edit.date().toString(Qt.DateFormat.ISODate),
            "Calibration Due": self.calibration_due_edit.date().toString(Qt.DateFormat.ISODate),
            "Unit Number": self.unit_number_edit.text(),
            "CustomerCompany": self.customer_edit.text(),
            "PhoneNumber": self.phone_edit.text(),
            "Address": self.address_edit.text(),
            "MaxTorque": f"{self.selected_row.get('max_torque', '')} {self.selected_row.get('unit', '')}" if self.selected_row else ""
        }
        if not summary_data:
            QMessageBox.warning(self, "Export Warning", "No table data to export.")
            return
        excel_save_dir = get_app_setting("excel_save_dir") or os.getcwd()
        pdf_save_dir = get_app_setting("pdf_save_dir") or os.getcwd()
        excel_filename_template = get_app_setting("excel_filename_template") or "summary_{{CustomerCompany}}_{{CalibrationDate}}.xlsx"
        pdf_filename_template = get_app_setting("pdf_filename_template") or "summary_{{CustomerCompany}}_{{CalibrationDate}}.pdf"
        filename_variables = {
            "Manufacturer": extra_info.get("Manufacturer", ""),
            "SerialNumber": extra_info.get("Serial Number", ""),
            "Model": extra_info.get("Model", ""),
            "CalibrationDate": extra_info.get("Calibration Date", ""),
            "CalibrationDue": extra_info.get("Calibration Due", ""),
            "UnitNumber": extra_info.get("Unit Number", ""),
            "CustomerCompany": extra_info.get("CustomerCompany", ""),
            "PhoneNumber": extra_info.get("PhoneNumber", ""),
            "Address": extra_info.get("Address", ""),
            "MaxTorque": extra_info.get("MaxTorque", "")
        }
        template_path = get_app_setting("summary_template_path") or "summary_template.xlsx"
        excel_path = None
        if self.excel_checkbox.isChecked():
            excel_filename = generate_filename(excel_filename_template, filename_variables)
            excel_path = os.path.join(excel_save_dir, excel_filename)
            try:
                if os.path.exists(template_path):
                    self.export_summary_with_template(template_path, extra_info, summary_data, excel_path)
                else:
                    df = pd.DataFrame(summary_data)
                    df.to_excel(excel_path, index=False)
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Error exporting Excel summary:\n{e}")
                return
        else:
            excel_path = None
        pdf_path = None
        if self.pdf_checkbox.isChecked():
            if not excel_path:
                QMessageBox.warning(self, "Export Warning", "PDF export requires Excel export to be enabled.")
                return
            pdf_filename = generate_filename(pdf_filename_template, filename_variables)
            pdf_path = os.path.join(pdf_save_dir, pdf_filename)
            try:
                convert_excel_to_pdf(excel_path, pdf_path)
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Error exporting PDF summary:\n{e}")
                return
        msg = "Summary exported to:\n"
        if excel_path:
            msg += f"Excel: {excel_path}\n"
        if pdf_path:
            msg += f"PDF: {pdf_path}"
        QMessageBox.information(self, "Export Summary", msg)
        self.last_exported_summary_path = excel_path if excel_path else pdf_path

    def export_summary_with_template(self, template_path, extra_info, summary_data, output_path):
        wb = load_workbook(template_path)
        ws = wb.active
        variables = {
            "Manufacturer": extra_info.get("Manufacturer", ""),
            "SerialNumber": extra_info.get("Serial Number", ""),
            "Model": extra_info.get("Model", ""),
            "CalibrationDate": extra_info.get("Calibration Date", ""),
            "CalibrationDue": extra_info.get("Calibration Due", ""),
            "UnitNumber": extra_info.get("Unit Number", ""),
            "CustomerCompany": extra_info.get("CustomerCompany", ""),
            "PhoneNumber": extra_info.get("PhoneNumber", ""),
            "Address": extra_info.get("Address", ""),
            "MaxTorque": extra_info.get("MaxTorque", "")
        }
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for key, val in variables.items():
                        placeholder = "{{" + key + "}}"
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(placeholder, str(val))
        summary_variables = {}
        for idx, row_data in enumerate(summary_data):
            allowance_number = idx + 1
            summary_variables[f"AppliedTorque{allowance_number}"] = row_data.get("Applied Torque", "")
            summary_variables[f"MinMaxAllowance{allowance_number}"] = row_data.get("Min - Max Allowance", "")
            for test in range(1, 6):
                summary_variables[f"Test{test}_Allowance{allowance_number}"] = row_data.get(f"Test {test}", "")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for key, val in summary_variables.items():
                        placeholder = "{{" + key + "}}"
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(placeholder, str(val))
        wb.save(output_path)

    # ----------------------------- EXPORTING ENVELOPE -----------------------------
    def export_envelope(self):
        row_count = self.torque_table.rowCount()
        headers = ["Applied Torque", "Min - Max Allowance", "Test 1", "Test 2", "Test 3", "Test 4", "Test 5"]
        summary_data = []
        for r in range(row_count):
            row_dict = {}
            for c in range(self.torque_table.columnCount()):
                item = self.torque_table.item(r, c)
                row_dict[headers[c]] = item.text() if item else ""
            summary_data.append(row_dict)
        extra_info = {
            "Manufacturer": self.manufacturer_edit.text(),
            "Serial Number": self.serial_number_edit.text(),
            "Model": self.model_edit.text(),
            "Calibration Date": self.calibration_date_edit.date().toString(Qt.DateFormat.ISODate),
            "Calibration Due": self.calibration_due_edit.date().toString(Qt.DateFormat.ISODate),
            "Unit Number": self.unit_number_edit.text(),
            "CustomerCompany": self.customer_edit.text(),
            "PhoneNumber": self.phone_edit.text(),
            "Address": self.address_edit.text(),
            "MaxTorque": f"{self.selected_row.get('max_torque', '')} {self.selected_row.get('unit', '')}" if self.selected_row else ""
        }
        if not summary_data:
            QMessageBox.warning(self, "Export Warning", "No table data to export.")
            return
        excel_save_dir = get_app_setting("excel_save_dir") or os.getcwd()
        pdf_save_dir = get_app_setting("pdf_save_dir") or os.getcwd()
        envelope_excel_filename_template = get_app_setting("envelope_excel_filename_template") or "envelope_{{CustomerCompany}}_{{CalibrationDate}}.xlsx"
        envelope_pdf_filename_template = get_app_setting("envelope_pdf_filename_template") or "envelope_{{CustomerCompany}}_{{CalibrationDate}}.pdf"
        filename_variables = {
            "Manufacturer": extra_info.get("Manufacturer", ""),
            "SerialNumber": extra_info.get("Serial Number", ""),
            "Model": extra_info.get("Model", ""),
            "CalibrationDate": extra_info.get("Calibration Date", ""),
            "CalibrationDue": extra_info.get("Calibration Due", ""),
            "UnitNumber": extra_info.get("Unit Number", ""),
            "CustomerCompany": extra_info.get("CustomerCompany", ""),
            "PhoneNumber": extra_info.get("PhoneNumber", ""),
            "Address": extra_info.get("Address", ""),
            "MaxTorque": extra_info.get("MaxTorque", "")
        }
        envelope_template_path = get_app_setting("envelope_template_path") or "envelope_template.xlsx"
        envelope_excel_path = None
        if self.envelope_excel_checkbox.isChecked():
            envelope_excel_filename = generate_filename(envelope_excel_filename_template, filename_variables)
            envelope_excel_path = os.path.join(excel_save_dir, envelope_excel_filename)
            try:
                if os.path.exists(envelope_template_path):
                    self.export_summary_with_template(envelope_template_path, extra_info, summary_data, envelope_excel_path)
                else:
                    df = pd.DataFrame(summary_data)
                    df.to_excel(envelope_excel_path, index=False)
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Error exporting Envelope Excel summary:\n{e}")
                return
        else:
            envelope_excel_path = None
        envelope_pdf_path = None
        if self.envelope_pdf_checkbox.isChecked():
            if not envelope_excel_path:
                QMessageBox.warning(self, "Export Warning", "Envelope PDF export requires Envelope Excel export to be enabled.")
                return
            envelope_pdf_filename = generate_filename(envelope_pdf_filename_template, filename_variables)
            envelope_pdf_path = os.path.join(pdf_save_dir, envelope_pdf_filename)
            try:
                convert_excel_to_pdf(envelope_excel_path, envelope_pdf_path)
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Error exporting Envelope PDF summary:\n{e}")
                return
        msg = "Envelope exported to:\n"
        if envelope_excel_path:
            msg += f"Excel: {envelope_excel_path}\n"
        if envelope_pdf_path:
            msg += f"PDF: {envelope_pdf_path}"
        QMessageBox.information(self, "Export Envelope", msg)
        self.last_exported_envelope_path = envelope_excel_path if envelope_excel_path else envelope_pdf_path

    # --------------------------- PRINTING FUNCTIONS ---------------------------
    def print_summary(self):
        if not self.last_exported_summary_path:
            QMessageBox.warning(self, "Print Warning", "No summary file available to print.")
            return
        try:
            os.startfile(self.last_exported_summary_path, "print")
            QMessageBox.information(self, "Print Summary", f"Summary sent to printer:\n{self.last_exported_summary_path}")
        except Exception as e:
            QMessageBox.critical(self, "Print Error", f"Error printing summary: {e}")

    def print_envelope(self):
        if not self.last_exported_envelope_path:
            QMessageBox.warning(self, "Print Warning", "No envelope file available to print.")
            return
        try:
            os.startfile(self.last_exported_envelope_path, "print")
            QMessageBox.information(self, "Print Envelope", f"Envelope sent to printer:\n{self.last_exported_envelope_path}")
        except Exception as e:
            QMessageBox.critical(self, "Print Error", f"Error printing envelope: {e}")

    # ------------------------------ SETTINGS TAB ------------------------------
    def init_settings_tab(self):
        self.settings_tab = QWidget()
        layout = QVBoxLayout(self.settings_tab)

        self.settings_combo = QComboBox()
        self.settings_combo.addItem("Data Management")
        self.settings_combo.addItem("OpenAI Settings")
        self.settings_combo.addItem("Export Settings")
        self.settings_combo.addItem("API Settings")
        self.settings_combo.addItem("Template Settings")
        self.settings_combo.addItem("Unit Synonyms")
        self.settings_combo.currentIndexChanged.connect(self.on_settings_combo_changed)
        layout.addWidget(self.settings_combo)

        self.settings_stacked = QStackedWidget()
        layout.addWidget(self.settings_stacked)

        # Data Management Page
        self.data_management_page = QWidget()
        dm_layout = QVBoxLayout(self.data_management_page)
        self.torque_table_widget = QTableWidget()
        self.torque_table_widget.setColumnCount(4)
        self.torque_table_widget.setHorizontalHeaderLabels(["Max Torque", "Unit", "Type", "Applied Torque"])
        self.torque_table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        dm_layout.addWidget(self.torque_table_widget)
        btn_layout = QHBoxLayout()
        self.add_btn = QPushButton("Add Entry")
        self.add_btn.clicked.connect(self.add_entry)
        self.edit_btn = QPushButton("Edit Entry")
        self.edit_btn.clicked.connect(self.edit_entry)
        self.delete_btn = QPushButton("Delete Entry")
        self.delete_btn.clicked.connect(self.delete_entry)
        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.clicked.connect(self.load_torque_table_data)
        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.edit_btn)
        btn_layout.addWidget(self.delete_btn)
        btn_layout.addWidget(self.refresh_btn)
        dm_layout.addLayout(btn_layout)
        self.extracted_data_checkbox = QCheckBox("Show Extracted Data in Testing Tab")
        self.extracted_data_checkbox.setChecked(self.show_extracted_data)
        self.extracted_data_checkbox.stateChanged.connect(self.toggle_extracted_data)
        dm_layout.addWidget(self.extracted_data_checkbox)
        self.data_management_page.setLayout(dm_layout)
        self.settings_stacked.addWidget(self.data_management_page)

        # OpenAI Settings Page
        self.openai_settings_page = QWidget()
        openai_layout = QFormLayout(self.openai_settings_page)
        self.api_key_edit = QLineEdit()
        if self.openai_api_key:
            self.api_key_edit.setText(self.openai_api_key)
        openai_layout.addRow("OpenAI API Key:", self.api_key_edit)
        self.model_combo = QComboBox()
        self.model_combo.addItems(["gpt-4o", "gpt-4o-mini", "gpt-4-turbo"])
        self.model_combo.setCurrentText(self.openai_model)
        openai_layout.addRow("Model:", self.model_combo)
        self.temp_spin = QDoubleSpinBox()
        self.temp_spin.setRange(0.0, 2.0)
        self.temp_spin.setSingleStep(0.1)
        self.temp_spin.setValue(self.openai_temperature)
        openai_layout.addRow("Temperature:", self.temp_spin)
        self.top_p_spin = QDoubleSpinBox()
        self.top_p_spin.setRange(0.0, 1.0)
        self.top_p_spin.setSingleStep(0.1)
        self.top_p_spin.setValue(self.openai_top_p)
        openai_layout.addRow("Top P:", self.top_p_spin)
        self.presence_spin = QDoubleSpinBox()
        self.presence_spin.setRange(0.0, 2.0)
        self.presence_spin.setSingleStep(0.1)
        self.presence_spin.setValue(self.openai_presence_penalty)
        openai_layout.addRow("Presence Penalty:", self.presence_spin)
        self.freq_spin = QDoubleSpinBox()
        self.freq_spin.setRange(0.0, 2.0)
        self.freq_spin.setSingleStep(0.1)
        self.freq_spin.setValue(self.openai_frequency_penalty)
        openai_layout.addRow("Frequency Penalty:", self.freq_spin)
        save_key_btn = QPushButton("Save OpenAI Settings")
        save_key_btn.clicked.connect(self.save_openai_settings)
        openai_layout.addWidget(save_key_btn)
        self.settings_stacked.addWidget(self.openai_settings_page)

        # Export Settings Page
        self.export_settings_page = QWidget()
        export_layout = QFormLayout(self.export_settings_page)
        self.excel_checkbox = QCheckBox("Enable Excel Export")
        self.excel_checkbox.setChecked(True)
        export_layout.addRow("", self.excel_checkbox)
        self.pdf_checkbox = QCheckBox("Enable PDF Export")
        self.pdf_checkbox.setChecked(True)
        export_layout.addRow("", self.pdf_checkbox)
        self.envelope_excel_checkbox = QCheckBox("Enable Envelope Excel Export")
        self.envelope_excel_checkbox.setChecked(True)
        export_layout.addRow("", self.envelope_excel_checkbox)
        self.envelope_pdf_checkbox = QCheckBox("Enable Envelope PDF Export")
        self.envelope_pdf_checkbox.setChecked(True)
        export_layout.addRow("", self.envelope_pdf_checkbox)
        self.excel_dir_edit = QLineEdit(get_app_setting("excel_save_dir") or os.getcwd())
        excel_dir_browse_btn = QPushButton("Browse")
        excel_dir_browse_btn.clicked.connect(self.browse_excel_dir)
        excel_dir_layout = QHBoxLayout()
        excel_dir_layout.addWidget(self.excel_dir_edit)
        excel_dir_layout.addWidget(excel_dir_browse_btn)
        export_layout.addRow("Excel Save Directory:", excel_dir_layout)
        self.pdf_dir_edit = QLineEdit(get_app_setting("pdf_save_dir") or os.getcwd())
        pdf_dir_browse_btn = QPushButton("Browse")
        pdf_dir_browse_btn.clicked.connect(self.browse_pdf_dir)
        pdf_dir_layout = QHBoxLayout()
        pdf_dir_layout.addWidget(self.pdf_dir_edit)
        pdf_dir_layout.addWidget(pdf_dir_browse_btn)
        export_layout.addRow("PDF Save Directory:", pdf_dir_layout)
        self.excel_template_edit = QLineEdit(get_app_setting("excel_filename_template") or "summary_{{CustomerCompany}}_{{CalibrationDate}}.xlsx")
        export_layout.addRow("Excel Filename Template:", self.excel_template_edit)
        self.pdf_template_edit = QLineEdit(get_app_setting("pdf_filename_template") or "summary_{{CustomerCompany}}_{{CalibrationDate}}.pdf")
        export_layout.addRow("PDF Filename Template:", self.pdf_template_edit)
        self.template_path_edit = QLineEdit(get_app_setting("summary_template_path") or "summary_template.xlsx")
        template_path_browse_btn = QPushButton("Browse")
        template_path_browse_btn.clicked.connect(self.browse_template_file)
        template_path_layout = QHBoxLayout()
        template_path_layout.addWidget(self.template_path_edit)
        template_path_layout.addWidget(template_path_browse_btn)
        export_layout.addRow("Summary Template File:", template_path_layout)
        # Envelope settings
        self.envelope_excel_template_edit = QLineEdit(get_app_setting("envelope_excel_filename_template") or "envelope_{{CustomerCompany}}_{{CalibrationDate}}.xlsx")
        export_layout.addRow("Envelope Excel Filename Template:", self.envelope_excel_template_edit)
        self.envelope_pdf_template_edit = QLineEdit(get_app_setting("envelope_pdf_filename_template") or "envelope_{{CustomerCompany}}_{{CalibrationDate}}.pdf")
        export_layout.addRow("Envelope PDF Filename Template:", self.envelope_pdf_template_edit)
        self.envelope_template_path_edit = QLineEdit(get_app_setting("envelope_template_path") or "envelope_template.xlsx")
        envelope_template_browse_btn = QPushButton("Browse")
        envelope_template_browse_btn.clicked.connect(self.browse_envelope_template_file)
        envelope_template_layout = QHBoxLayout()
        envelope_template_layout.addWidget(self.envelope_template_path_edit)
        envelope_template_layout.addWidget(envelope_template_browse_btn)
        export_layout.addRow("Envelope Template File:", envelope_template_layout)
        save_export_btn = QPushButton("Save Export Settings")
        save_export_btn.clicked.connect(self.save_export_settings)
        export_layout.addWidget(save_export_btn)
        self.settings_stacked.addWidget(self.export_settings_page)

        # API Settings Page
        self.api_settings_page = QWidget()
        api_layout = QFormLayout(self.api_settings_page)
        self.laravel_url_edit = QLineEdit(get_app_setting("laravel_app_url") or "https://dev.c-trac.app")
        api_layout.addRow("Laravel App URL:", self.laravel_url_edit)
        self.laravel_token_edit = QLineEdit(get_app_setting("laravel_api_token") or "")
        api_layout.addRow("Laravel API Token:", self.laravel_token_edit)
        save_api_btn = QPushButton("Save API Settings")
        save_api_btn.clicked.connect(self.save_api_settings)
        api_layout.addWidget(save_api_btn)
        self.settings_stacked.addWidget(self.api_settings_page)

        # Template Settings Page
        self.template_settings_page = QWidget()
        template_set_layout = QFormLayout(self.template_settings_page)
        self.base_template_path_edit = QLineEdit(get_app_setting("base_template_path") or os.getcwd())
        template_set_layout.addRow("Base Template Save Path:", self.base_template_path_edit)
        create_template_btn = QPushButton("Create Base Template")
        create_template_btn.clicked.connect(self.create_base_template_action)
        template_set_layout.addWidget(create_template_btn)
        self.template_settings_page.setLayout(template_set_layout)
        self.settings_stacked.addWidget(self.template_settings_page)

        # Unit Synonyms Settings Page
        self.unit_synonyms_page = QWidget()
        unit_synonyms_layout = QFormLayout(self.unit_synonyms_page)
        self.ft_lb_synonyms_edit = QLineEdit(get_app_setting("synonyms_ft_lb") or "ft/lb,ft-lb,ft.lb,ft lb,ft/lbs,ft-lbs,ft.lbs,ft lbs")
        unit_synonyms_layout.addRow("FT/LB Synonyms:", self.ft_lb_synonyms_edit)
        self.in_lb_synonyms_edit = QLineEdit(get_app_setting("synonyms_in_lb") or "in/lb,in-lb,in.lb,in lb,in/lbs,in-lbs,in.lbs,in lbs")
        unit_synonyms_layout.addRow("IN/LB Synonyms:", self.in_lb_synonyms_edit)
        self.nm_synonyms_edit = QLineEdit(get_app_setting("synonyms_nm") or "nm,n.m,n*m,nm.n.m.")
        unit_synonyms_layout.addRow("NM Synonyms:", self.nm_synonyms_edit)
        save_unit_synonyms_btn = QPushButton("Save Unit Synonyms")
        save_unit_synonyms_btn.clicked.connect(self.save_unit_synonyms)
        unit_synonyms_layout.addWidget(save_unit_synonyms_btn)
        self.unit_synonyms_page.setLayout(unit_synonyms_layout)
        self.settings_stacked.addWidget(self.unit_synonyms_page)

        self.settings_tab.setLayout(layout)
        self.tab_widget.addTab(self.settings_tab, "Settings")

    def on_settings_combo_changed(self, index):
        self.settings_stacked.setCurrentIndex(index)

    def browse_excel_dir(self):
        directory = QFileDialog.getExistingDirectory(self, "Select Excel Save Directory")
        if directory:
            self.excel_dir_edit.setText(directory)

    def browse_pdf_dir(self):
        directory = QFileDialog.getExistingDirectory(self, "Select PDF Save Directory")
        if directory:
            self.pdf_dir_edit.setText(directory)

    def browse_template_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Summary Template", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_path:
            self.template_path_edit.setText(file_path)

    def browse_envelope_template_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Envelope Template", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_path:
            self.envelope_template_path_edit.setText(file_path)

    def save_export_settings(self):
        set_app_setting("excel_save_dir", self.excel_dir_edit.text())
        set_app_setting("pdf_save_dir", self.pdf_dir_edit.text())
        set_app_setting("excel_filename_template", self.excel_template_edit.text())
        set_app_setting("pdf_filename_template", self.pdf_template_edit.text())
        set_app_setting("summary_template_path", self.template_path_edit.text())
        set_app_setting("envelope_excel_filename_template", self.envelope_excel_template_edit.text())
        set_app_setting("envelope_pdf_filename_template", self.envelope_pdf_template_edit.text())
        set_app_setting("envelope_template_path", self.envelope_template_path_edit.text())
        QMessageBox.information(self, "Settings Saved", "Export settings have been saved.")

    def save_api_settings(self):
        set_app_setting("laravel_app_url", self.laravel_url_edit.text())
        set_app_setting("laravel_api_token", self.laravel_token_edit.text())
        QMessageBox.information(self, "Settings Saved", "API settings have been saved.")

    def save_unit_synonyms(self):
        set_app_setting("synonyms_ft_lb", self.ft_lb_synonyms_edit.text())
        set_app_setting("synonyms_in_lb", self.in_lb_synonyms_edit.text())
        set_app_setting("synonyms_nm", self.nm_synonyms_edit.text())
        QMessageBox.information(self, "Settings Saved", "Unit synonyms have been saved.")

    def create_base_template_action(self):
        path = self.base_template_path_edit.text().strip()
        if not path:
            QMessageBox.warning(self, "Invalid Path", "Please enter a valid save path for the base template.")
            return
        filename = os.path.join(path, "base_template.xlsx")
        if self.create_base_template(filename):
            set_app_setting("base_template_path", path)
            QMessageBox.information(self, "Template Created", f"Base template created at:\n{filename}")

    def create_base_template(self, filename):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Base Template"
            ws["A1"] = "Torque Test Report Template"
            ws["A2"] = "Manufacturer: {{Manufacturer}}"
            ws["A3"] = "Model: {{Model}}"
            ws["A4"] = "Unit Number: {{UnitNumber}}"
            ws["A5"] = "Serial Number: {{SerialNumber}}"
            ws["A6"] = "Customer: {{CustomerCompany}}"
            ws["A7"] = "Phone: {{PhoneNumber}}"
            ws["A8"] = "Address: {{Address}}"
            ws["A9"] = "Max Torque: {{MaxTorque}}"
            ws["A10"] = "Calibration Date: {{CalibrationDate}}"
            ws["A11"] = "Calibration Due: {{CalibrationDue}}"
            wb.save(filename)
            return True
        except Exception as e:
            QMessageBox.critical(self, "Template Creation Error", f"Error creating base template:\n{e}")
            return False

    def load_torque_table_data(self):
        table_data = get_torque_table()
        self.torque_table_widget.setRowCount(len(table_data))
        for i, row in enumerate(table_data):
            self.torque_table_widget.setItem(i, 0, QTableWidgetItem(str(row.get("max_torque", ""))))
            self.torque_table_widget.setItem(i, 1, QTableWidgetItem(row.get("unit", "")))
            self.torque_table_widget.setItem(i, 2, QTableWidgetItem(row.get("type", "")))
            self.torque_table_widget.setItem(i, 3, QTableWidgetItem(row.get("applied_torq", "")))

    def add_entry(self):
        dialog = TorqueEntryDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            add_torque_entry(
                data["max_torque"], data["unit"], data["type"],
                data["applied_torq"], data["allowance1"], data["allowance2"], data["allowance3"]
            )
            self.load_max_torque_dropdown()
            self.load_torque_table_data()

    def edit_entry(self):
        selected_items = self.torque_table_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Edit Entry", "No entry selected.")
            return
        row = selected_items[0].row()
        table_data = get_torque_table()
        if row >= len(table_data):
            return
        entry = table_data[row]
        dialog = TorqueEntryDialog(self, entry)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            update_torque_entry(
                entry["id"], data["max_torque"], data["unit"], data["type"],
                data["applied_torq"], data["allowance1"], data["allowance2"], data["allowance3"]
            )
            self.load_max_torque_dropdown()
            self.load_torque_table_data()

    def delete_entry(self):
        selected_items = self.torque_table_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Delete Entry", "No entry selected.")
            return
        row = selected_items[0].row()
        table_data = get_torque_table()
        if row >= len(table_data):
            return
        entry = table_data[row]
        delete_torque_entry(entry["id"])
        self.load_max_torque_dropdown()
        self.load_torque_table_data()

    def toggle_extracted_data(self, state):
        self.show_extracted_data = (state == Qt.CheckState.Checked)
        self.extracted_data_label.setVisible(self.show_extracted_data)
        self.extracted_data_table.setVisible(self.show_extracted_data)

    # ------------------------- REPORT TEMPLATES TAB -------------------------
    def init_report_templates_tab(self):
        # Restore original report templates functionality as in your uploaded file
        report_tab = QWidget()
        layout = QVBoxLayout(report_tab)
        layout.addWidget(QLabel("Report Templates Functionality:"))
        # (Implement additional report template creation, preview, or editing functions here as needed.)
        layout.addWidget(QLabel("Use the settings under the Export and Template Settings pages to configure report templates."))
        report_tab.setLayout(layout)
        self.tab_widget.addTab(report_tab, "Report Templates")

    # ------------------------------ SETTINGS SAVE METHODS ------------------------------
    def save_openai_settings(self):
        set_app_setting("openai_api_key", self.api_key_edit.text())
        set_app_setting("openai_model", self.model_combo.currentText())
        set_app_setting("openai_temperature", str(self.temp_spin.value()))
        set_app_setting("openai_top_p", str(self.top_p_spin.value()))
        set_app_setting("openai_presence_penalty", str(self.presence_spin.value()))
        set_app_setting("openai_frequency_penalty", str(self.freq_spin.value()))
        QMessageBox.information(self, "Settings Saved", "OpenAI settings have been saved.")

    def save_api_settings(self):
        set_app_setting("laravel_app_url", self.laravel_url_edit.text())
        set_app_setting("laravel_api_token", self.laravel_token_edit.text())
        QMessageBox.information(self, "Settings Saved", "API settings have been saved.")

    # ---------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    from PyQt6.QtWidgets import QApplication
    init_db()
    insert_default_torque_table_data()
    app = QApplication(sys.argv)
    window = ModernTorqueApp()
    window.show()
    sys.exit(app.exec())
